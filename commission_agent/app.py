import streamlit as st
import pdfplumber
import openpyxl
import json
import io
import re
import pandas as pd
from pathlib import Path
from groq import Groq

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Commission Agent — MAR1",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stSidebar"] { background-color: #181c26; }
[data-testid="stSidebar"] * { color: #e8eaf0 !important; }
.main-header {
    background: linear-gradient(135deg, #4f8ef7, #7c3aed);
    padding: 18px 24px; border-radius: 12px; margin-bottom: 24px;
}
.main-header h1 { color: white; margin: 0; font-size: 22px; }
.main-header p  { color: rgba(255,255,255,0.75); margin: 0; font-size: 13px; }
.ok-box   { background: rgba(34,197,94,0.08);  border: 1px solid rgba(34,197,94,0.2);  border-radius:10px; padding:12px 16px; margin-bottom:8px; }
.err-box  { background: rgba(239,68,68,0.08);  border: 1px solid rgba(239,68,68,0.25); border-radius:10px; padding:14px 16px; margin-bottom:10px; }
</style>
""", unsafe_allow_html=True)

# ─── All 79 header columns (exact order from Header.xlsx) ─────────────────────
HEADER_COLS = [
    'Distname', 'Supplier_name', 'direct_indirect', 'in_out_territory',
    'CustAccNbr', 'CustDunsID', 'CustName', 'Address1', 'City', 'State',
    'County', 'Zip', 'Phone', 'Country', 'NoOfEmployees', 'WebAddress',
    'SIC', 'NAICS', 'LineOfBusiness', 'ParentName', 'AccountType', 'UOM',
    'InvoiceNumber', 'Qty', 'UnitCost', 'UnitResale', 'InvoiceDate',
    'DateRecieved', 'PartNumberSubmitted', 'PartNumberDescription', 'Branch',
    'SalesRep', 'Latitude', 'Longitude', 'Brand', 'PartNumberActual',
    'UPCCode', 'rawcustname', 'rawdistaddress', 'rawdistcity', 'rawdiststate',
    'rawdistpostalcode', 'rawdistcountry', 'currency', 'contractID',
    'client_CustName', 'Zip_4_digit', 'dnb_trade_style', 'dnb_sales_value',
    'google_CustName', 'google_Address1', 'google_State', 'google_Zip',
    'google_Country', 'google_Phone', 'google_WebAddress', 'Pay_Month',
    'Pay_Year', 'Ship_Month', 'Ship_Year', 'Industry', 'Commissions',
    'Commission_Rate', 'Cust_AM', 'CEM', 'Sales', 'In_Out',
    'Commission_split_percentage', 'Distributor_part_number', 'Category',
    'google_City', 'Billings', 'Cheque_Number', 'Pay_Date', 'meta_data_json',
    'SO_Number', 'PO_Number', 'ship_date', 'searched_on_google'
]

# ─── Session State ─────────────────────────────────────────────────────────────
if "analysis"     not in st.session_state: st.session_state.analysis     = None
if "chat_history" not in st.session_state: st.session_state.chat_history = []
if "api_key"      not in st.session_state:
    try:    st.session_state.api_key = st.secrets["GROQ_API_KEY"]
    except: st.session_state.api_key = ""

GROQ_MODEL = "llama-3.3-70b-versatile"

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt(n):
    try:    return f"${float(n):,.2f}"
    except: return "$0.00"

def fmtpct(n):
    try:    return f"{float(n)*100:.1f}%"
    except: return "—"

def extract_pdf_text(file_bytes: bytes) -> str:
    text = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text.append(t)
    return "\n".join(text)

def extract_xlsx_sheets(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[sheet] = rows
    return sheets

def sheets_to_text(sheets: dict) -> str:
    lines = []
    for name, rows in sheets.items():
        lines.append(f"=== Sheet: {name} ===")
        for row in rows:
            vals = [str(v) if v is not None else "" for v in row]
            if any(v.strip() for v in vals):
                lines.append("\t".join(vals))
    return "\n".join(lines)

# ─── Map extracted data → Header columns ──────────────────────────────────────
def map_to_header(data: dict) -> pd.DataFrame:
    """
    Map Part I (PDF) and Part II (Excel INTEGRA) rows
    into all 79 header columns from Header.xlsx
    """
    all_rows = []

    p1_month = data.get("part1", {}).get("month", "")
    p2_month = data.get("part2", {}).get("month", "")

    # Parse month/year helpers
    def get_month_year(label):
        """e.g. 'December 2025' → (12, 2025)"""
        try:
            from datetime import datetime
            dt = datetime.strptime(label, "%B %Y")
            return dt.month, dt.year
        except:
            return "", ""

    p1_mon, p1_yr = get_month_year(p1_month)
    p2_mon, p2_yr = get_month_year(p2_month)
    comm_rate      = data.get("part2", {}).get("rate", 0.05)

    # ── Part I: Regular Sales (from PDF) ──────────────────────────────────────
    for row in data.get("part1", {}).get("rows", []):
        r = {col: "" for col in HEADER_COLS}
        # Identity
        r["Distname"]          = "MARCTECH2, INC."
        r["Supplier_name"]     = "AMERICAN BRIGHT OPTOELECTRONICS CORP."
        r["direct_indirect"]   = "Direct"
        r["in_out_territory"]  = "In"
        r["SalesRep"]          = "MAR1"
        r["Branch"]            = "MAR1"
        r["currency"]          = "USD"
        # Customer
        r["CustAccNbr"]        = row.get("customer_number", "")
        r["CustName"]          = row.get("customer_name", "")
        r["State"]             = row.get("state", "")
        r["Country"]           = "US"
        # Invoice
        r["InvoiceNumber"]     = row.get("invoice_number", "")
        r["InvoiceDate"]       = row.get("invoice_date", "")
        r["Qty"]               = 1
        r["UnitCost"]          = row.get("unit_cost", "")      # Sales Amt ✅
        r["UnitResale"]        = ""                             # N/A for direct
        r["Commissions"]       = row.get("commission", "")
        r["Commission_Rate"]   = 0.05
        r["Billings"]          = row.get("unit_cost", "")
        # PO / Part
        r["PO_Number"]         = row.get("po_number", "")
        r["PartNumberSubmitted"] = row.get("po_number", "")
        # Date fields
        r["Pay_Month"]         = p1_mon
        r["Pay_Year"]          = p1_yr
        r["Ship_Month"]        = p1_mon
        r["Ship_Year"]         = p1_yr
        # Raw / client / google mirrors
        r["rawcustname"]       = row.get("customer_name", "")
        r["client_CustName"]   = row.get("customer_name", "")
        r["google_CustName"]   = row.get("customer_name", "")
        r["google_State"]      = row.get("state", "")
        r["google_Country"]    = "US"
        r["In_Out"]            = "In"
        all_rows.append(r)

    # ── Part II: Distributor Sales (from Excel INTEGRA) ───────────────────────
    for row in data.get("part2", {}).get("rows", []):
        r = {col: "" for col in HEADER_COLS}
        # Identity
        r["Distname"]          = "MARCTECH2, INC."
        r["Supplier_name"]     = "AMERICAN BRIGHT OPTOELECTRONICS CORP."
        r["direct_indirect"]   = "Indirect"
        r["in_out_territory"]  = "In"
        r["SalesRep"]          = "MAR1"
        r["Branch"]            = "MAR1"
        r["currency"]          = "USD"
        # Customer
        r["CustName"]          = row.get("customer_name", "")
        r["City"]              = row.get("city", "")
        r["State"]             = row.get("state", "")
        r["Zip"]               = row.get("zip", "")
        r["Country"]           = "US"
        # Invoice
        r["InvoiceNumber"]     = row.get("invoice_number", "")
        r["Qty"]               = row.get("quantity", "")
        r["UnitCost"]          = ""                             # N/A for disty
        r["UnitResale"]        = row.get("unit_sale", "")      # AB PRICE ✅
        r["Commissions"]       = row.get("commission", "")
        r["Commission_Rate"]   = comm_rate
        r["Billings"]          = row.get("total_amount", "")
        # Part number
        r["PartNumberSubmitted"]   = row.get("part_number", "")
        r["PartNumberActual"]      = row.get("part_number", "")
        r["Distributor_part_number"] = row.get("part_number", "")
        # Date fields
        r["Pay_Month"]         = p1_mon
        r["Pay_Year"]          = p1_yr
        r["Ship_Month"]        = p2_mon
        r["Ship_Year"]         = p2_yr
        # Raw / client / google mirrors
        r["rawcustname"]       = row.get("customer_name", "")
        r["rawdistcity"]       = row.get("city", "")
        r["rawdiststate"]      = row.get("state", "")
        r["rawdistpostalcode"] = row.get("zip", "")
        r["client_CustName"]   = row.get("customer_name", "")
        r["google_CustName"]   = row.get("customer_name", "")
        r["google_City"]       = row.get("city", "")
        r["google_State"]      = row.get("state", "")
        r["google_Zip"]        = row.get("zip", "")
        r["google_Country"]    = "US"
        r["In_Out"]            = "In"
        all_rows.append(r)

    df = pd.DataFrame(all_rows, columns=HEADER_COLS)
    return df

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    return buf.getvalue()

# ─── Groq Analysis ────────────────────────────────────────────────────────────
def run_analysis(pdf_text: str, xlsx_sheets: dict, api_key: str) -> dict:
    client = Groq(api_key=api_key)
    xlsx_text = sheets_to_text(xlsx_sheets)

    prompt = f"""
=== PDF FILE: Sales Commission Report from Payment ===
{pdf_text}

=== EXCEL FILE (all sheets) ===
{xlsx_text}

---
You are a commission data extraction agent for MARCTECH2, INC. (Sales Rep MAR1).

TASK: Extract structured data using these EXACT rules:

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PART I — REGULAR SALES (from PDF)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
For each transaction row in the PDF extract:
  - invoice_number   ← "Invoice#" column
  - invoice_date     ← "Inv. Dt" column
  - po_number        ← "PO #" column
  - customer_number  ← "Customer#" column
  - customer_name    ← "Name" column
  - unit_cost        ← "Sales Amt" column (dollar amount only, NOT the % figure)
  - state            ← "State" column
  - commission       ← "Commission" column (dollar amount only, NOT the 5.00% figure)

Also extract:
  - total_amount       ← Sales Rep Total payment amount
  - total_commission   ← Sales Rep Total commission amount
  - month_label        ← e.g. "December 2025"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PART II — DISTRIBUTOR SALES (from Excel)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Step 1 — Get commission RATE from DISTY SALES sheet (RATE column per AB INV NO.)
Step 2 — Get line items from INTEGRA sheet:
  - state          ← "ST"
  - zip            ← "Zip"
  - part_number    ← "Item No."
  - quantity       ← "Qty"
  - unit_sale      ← "AB PRICE"   ← this goes into UnitResale in the output
  - invoice_number ← "AB. Inv. No."
  - customer_name  ← "Company"
  - city           ← "City"
Step 3 — Calculate:
  - total_amount = quantity × unit_sale
  - commission   = total_amount × rate

Also extract:
  - part2_month_label      ← month from DISTY SALES header
  - part2_total_amount
  - part2_total_commission
  - rate                   ← the commission rate (e.g. 0.05)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COMM REPORT SUMMARY (from COMM REPORT sheet)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  - comm_report_part1_amount
  - comm_report_part1_commission
  - comm_report_part2_amount
  - comm_report_part2_commission
  - comm_report_total_commission

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RECONCILIATION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1. Does PDF Part I commission = COMM REPORT Part I commission? (tolerance $0.01)
  2. Does calculated Part II commission = COMM REPORT Part II commission? (tolerance $0.01)

Return this EXACT JSON (no markdown, no explanation):

{{
  "month_label": "December 2025",
  "part1": {{
    "month": "December 2025",
    "total_amount": 6090.00,
    "total_commission": 304.50,
    "rows": [
      {{
        "invoice_number": "451014081",
        "invoice_date": "11/12/25",
        "po_number": "5034774",
        "customer_number": "DEL003",
        "customer_name": "DELTA CONTROLS, INC.",
        "unit_cost": 6090.00,
        "state": "BC",
        "commission": 304.50
      }}
    ]
  }},
  "part2": {{
    "month": "November 2025",
    "rate": 0.05,
    "total_amount": 572.00,
    "total_commission": 28.60,
    "rows": [
      {{
        "invoice_number": "5034813",
        "customer_name": "SYMETRIX, INC.",
        "city": "MUKILTEO",
        "state": "WA",
        "zip": "98275",
        "part_number": "BA-10G1UD",
        "quantity": 1000,
        "unit_sale": 0.572,
        "total_amount": 572.00,
        "commission": 28.60
      }}
    ]
  }},
  "comm_report": {{
    "part1_amount": 6090.00,
    "part1_commission": 304.50,
    "part2_amount": 572.00,
    "part2_commission": 28.60,
    "total_commission": 333.10
  }},
  "reconciliation": {{
    "part1_match": true,
    "part2_match": true,
    "issues": [],
    "status": "ok"
  }}
}}
"""

    response = client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=4000,
        temperature=0.1,
    )
    raw = response.choices[0].message.content
    raw = re.sub(r"```json|```", "", raw).strip()
    m = re.search(r'\{.*\}', raw, re.DOTALL)
    if m: raw = m.group()
    return json.loads(raw)


def chat_with_agent(question, context, history, api_key):
    client = Groq(api_key=api_key)
    system = (
        "You are a commission analysis assistant for MARCTECH2, INC. (Sales Rep MAR1). "
        "Answer questions about commission data clearly. Use dollar amounts.\n\n"
        f"DATA:\n{json.dumps(context, indent=2)}"
    ) if context else "No data analyzed yet."
    messages = [{"role": "system", "content": system}]
    for h in history[-6:]:
        messages.append({"role": "user",      "content": h["user"]})
        messages.append({"role": "assistant", "content": h["assistant"]})
    messages.append({"role": "user", "content": question})
    resp = client.chat.completions.create(model=GROQ_MODEL, messages=messages, max_tokens=800, temperature=0.3)
    return resp.choices[0].message.content


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.markdown("Free API key → [console.groq.com](https://console.groq.com)")
    api_key = st.text_input("Groq API Key", type="password",
                             value=st.session_state.api_key, placeholder="gsk_…")
    if api_key:
        st.session_state.api_key = api_key

    st.markdown("---")
    st.markdown("## 📂 Upload Files")
    pdf_file  = st.file_uploader("PDF — Comm_from_Payment",  type=["pdf"],         key="pdf_up")
    xlsx_file = st.file_uploader("Excel — Comm_Report",       type=["xlsx", "xls"], key="xl_up")

    st.markdown("---")
    run = st.button("⚡ Analyze & Reconcile", use_container_width=True,
                    disabled=not (pdf_file and xlsx_file and st.session_state.api_key))
    if not st.session_state.api_key:
        st.caption("⚠️ Enter Groq API key above.")
    elif not (pdf_file and xlsx_file):
        st.caption("⚠️ Upload both PDF and Excel files.")

# ─── Run Analysis ─────────────────────────────────────────────────────────────
if run and pdf_file and xlsx_file:
    with st.spinner("Reading files…"):
        pdf_text   = extract_pdf_text(pdf_file.read())
        xlsx_sheets = extract_xlsx_sheets(xlsx_file.read())

    with st.spinner("Extracting & reconciling with AI…"):
        try:
            st.session_state.analysis = run_analysis(pdf_text, xlsx_sheets, st.session_state.api_key)
            st.success("Done!")
        except Exception as e:
            st.error(f"Failed: {e}")

# ─── Main UI ──────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📊 Commission Intelligence Agent</h1>
  <p>MARCTECH2, INC. · American Bright Optoelectronics Corp. · Sales Rep: MAR1 · Powered by Groq (Free)</p>
</div>
""", unsafe_allow_html=True)

data = st.session_state.analysis

if not data:
    st.info("👈 Upload the PDF and Excel files in the sidebar, then click **Analyze & Reconcile**.")
    with st.expander("📋 Field Mapping Reference"):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**PDF → Header columns (Part I)**")
            st.markdown("""
| PDF Column | Header Column |
|---|---|
| Invoice# | InvoiceNumber |
| Inv. Dt | InvoiceDate |
| PO # | PartNumberSubmitted |
| Customer# | CustAccNbr |
| Name | CustName |
| Sales Amt | UnitCost |
| State | State |
| Commission | *(calculated)* |
""")
        with col2:
            st.markdown("**Excel INTEGRA → Header columns (Part II)**")
            st.markdown("""
| Excel Column | Header Column |
|---|---|
| Company | CustName |
| City | City |
| ST | State |
| Zip | Zip |
| Item No. | PartNumberSubmitted / PartNumberActual |
| Qty | Qty |
| **AB PRICE** | **UnitResale** ✅ |
| AB. Inv. No. | InvoiceNumber |
""")
        st.info("Commission = Qty × AB PRICE × Rate (from DISTY SALES sheet)")
else:
    p1  = data.get("part1",       {})
    p2  = data.get("part2",       {})
    cr  = data.get("comm_report", {})
    rec = data.get("reconciliation", {})

    # Build the header-mapped DataFrame
    header_df = map_to_header(data)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Summary",
        "📄 Part I — Regular Sales",
        "🏪 Part II — Distributor Sales",
        "📋 Export (Header Format)",
        "💬 Ask AI",
    ])

    # ── Tab 1: Summary ───────────────────────────────────────────────────────
    with tab1:
        st.markdown(f"### Commission Statement — {data.get('month_label','')}")
        c1, c2, c3 = st.columns(3)
        c1.metric("Part I — Regular Sales",    fmt(cr.get("part1_amount", 0)),
                  f"Commission: {fmt(cr.get('part1_commission', 0))}")
        c2.metric("Part II — Distributor Sales", fmt(cr.get("part2_amount", 0)),
                  f"Commission: {fmt(cr.get('part2_commission', 0))}")
        c3.metric("💰 Total Commission",       fmt(cr.get("total_commission", 0)))

        st.markdown("---")
        st.markdown("### Reconciliation")
        status = rec.get("status", "ok")
        if status == "ok":
            st.markdown('<div class="ok-box">✅ <strong>All figures reconcile perfectly.</strong></div>',
                        unsafe_allow_html=True)
        else:
            for issue in rec.get("issues", []):
                st.markdown(f'<div class="err-box">⚠️ {issue}</div>', unsafe_allow_html=True)

        col_a, col_b, col_c, col_d = st.columns([3, 2, 2, 1])
        col_a.markdown("**Check**")
        col_b.markdown("**Calculated**")
        col_c.markdown("**COMM REPORT**")
        col_d.markdown("**Match**")
        col_a.write("Part I — PDF Commission")
        col_b.write(fmt(p1.get("total_commission")))
        col_c.write(fmt(cr.get("part1_commission")))
        col_d.write("✅" if rec.get("part1_match") else "❌")
        col_a.write("Part II — Distributor Commission")
        col_b.write(fmt(p2.get("total_commission")))
        col_c.write(fmt(cr.get("part2_commission")))
        col_d.write("✅" if rec.get("part2_match") else "❌")

    # ── Tab 2: Part I ────────────────────────────────────────────────────────
    with tab2:
        st.markdown(f"### Part I — Regular Sales &nbsp; `{p1.get('month','')}`")
        rows1 = p1.get("rows", [])
        if rows1:
            df1 = pd.DataFrame(rows1).rename(columns={
                "invoice_number":  "InvoiceNumber",
                "invoice_date":    "InvoiceDate",
                "po_number":       "PO #",
                "customer_number": "CustAccNbr",
                "customer_name":   "CustName",
                "unit_cost":       "UnitCost",
                "state":           "State",
                "commission":      "Commission",
            })
            for col in ["UnitCost", "Commission"]:
                if col in df1.columns:
                    df1[col] = df1[col].apply(lambda x: f"${float(x):,.2f}" if x else "—")
            st.dataframe(df1, use_container_width=True, hide_index=True)
        else:
            st.warning("No rows extracted.")
        c1, c2 = st.columns(2)
        c1.metric("Total Sales Amount",    fmt(p1.get("total_amount")))
        c2.metric("Total Commission (5%)", fmt(p1.get("total_commission")))

    # ── Tab 3: Part II ───────────────────────────────────────────────────────
    with tab3:
        st.markdown(f"### Part II — Distributor Sales &nbsp; `{p2.get('month','')}`")
        st.caption(f"Rate from DISTY SALES: **{fmtpct(p2.get('rate', 0.05))}** &nbsp;|&nbsp; Formula: Qty × AB PRICE × Rate = Commission")
        rows2 = p2.get("rows", [])
        if rows2:
            df2 = pd.DataFrame(rows2).rename(columns={
                "invoice_number": "InvoiceNumber",
                "customer_name":  "CustName",
                "city":           "City",
                "state":          "State",
                "zip":            "Zip",
                "part_number":    "PartNumber",
                "quantity":       "Qty",
                "unit_sale":      "UnitResale (AB PRICE)",
                "total_amount":   "Total Amount",
                "commission":     "Commission",
            })
            for col in ["UnitResale (AB PRICE)", "Total Amount", "Commission"]:
                if col in df2.columns:
                    df2[col] = df2[col].apply(lambda x: f"${float(x):,.4f}" if "PRICE" in col else f"${float(x):,.2f}" if x else "—")
            st.dataframe(df2, use_container_width=True, hide_index=True)
        else:
            st.warning("No rows extracted.")
        c1, c2, c3 = st.columns(3)
        c1.metric("Rate",                     fmtpct(p2.get("rate", 0.05)))
        c2.metric("Total Distributor Amount", fmt(p2.get("total_amount")))
        c3.metric("Total Commission",         fmt(p2.get("total_commission")))

    # ── Tab 4: Export ────────────────────────────────────────────────────────
    with tab4:
        st.markdown("### 📋 Data Mapped to Header Format")
        st.caption("All 50 columns from Header.xlsx — AB PRICE mapped to **UnitResale**, Sales Amt to **UnitCost**.")

        st.dataframe(header_df, use_container_width=True, hide_index=True)

        # Download button
        excel_bytes = to_excel_bytes(header_df)
        month_label = data.get("month_label", "output").replace(" ", "_")
        st.download_button(
            label="⬇️ Download as Excel (Header Format)",
            data=excel_bytes,
            file_name=f"Commission_{month_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.markdown("---")
        st.markdown("**Column mapping used (all 79 columns):**")
        mapping_data = {
            "Header Column": [
                "Distname","Supplier_name","direct_indirect","in_out_territory",
                "CustAccNbr","CustName","City","State","Zip","Country",
                "InvoiceNumber","InvoiceDate","Qty","UnitCost","UnitResale",
                "Commissions","Commission_Rate","Billings","PO_Number",
                "PartNumberSubmitted","PartNumberActual","Distributor_part_number",
                "SalesRep","Branch","currency","Pay_Month","Pay_Year",
                "Ship_Month","Ship_Year","In_Out",
                "rawcustname","rawdistcity","rawdiststate","rawdistpostalcode",
                "client_CustName","google_CustName","google_City","google_State","google_Zip","google_Country",
            ],
            "Part I — PDF Source": [
                "MARCTECH2, INC.","AMERICAN BRIGHT...","Direct","In",
                "Customer#","Name","—","State","—","US",
                "Invoice#","Inv. Dt","1 (direct)","Sales Amt ✅","—",
                "Commission ($)","0.05","Sales Amt","PO #",
                "PO #","—","—",
                "MAR1","MAR1","USD","from PDF month","from PDF month",
                "from PDF month","from PDF month","In",
                "Name","—","State","—",
                "Name","Name","—","State","—","US",
            ],
            "Part II — Excel Source": [
                "MARCTECH2, INC.","AMERICAN BRIGHT...","Indirect","In",
                "—","Company","City","ST","Zip","US",
                "AB. Inv. No.","—","Qty","—","AB PRICE ✅",
                "Qty×AB PRICE×Rate","Rate (DISTY SALES)","Qty×AB PRICE","—",
                "Item No.","Item No.","Item No.",
                "MAR1","MAR1","USD","from pay month","from pay month",
                "from disty month","from disty month","In",
                "Company","City","ST","Zip",
                "Company","Company","City","ST","Zip","US",
            ],
        }
        st.dataframe(pd.DataFrame(mapping_data), use_container_width=True, hide_index=True)

    # ── Tab 5: Ask AI ────────────────────────────────────────────────────────
    with tab5:
        st.markdown("Ask anything about this commission statement.")
        for h in st.session_state.chat_history:
            with st.chat_message("user"):      st.write(h["user"])
            with st.chat_message("assistant"): st.write(h["assistant"])

        question = st.chat_input("e.g. What is the total commission? Any discrepancies?")
        if question:
            if not st.session_state.api_key:
                st.error("Enter Groq API key in the sidebar.")
            else:
                with st.chat_message("user"): st.write(question)
                with st.chat_message("assistant"):
                    with st.spinner("Thinking…"):
                        reply = chat_with_agent(question, data, st.session_state.chat_history, st.session_state.api_key)
                    st.write(reply)
                st.session_state.chat_history.append({"user": question, "assistant": reply})
