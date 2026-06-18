import streamlit as st
import pdfplumber
import openpyxl
import anthropic
import json
import io
import re
from pathlib import Path

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Commission Agent — MAR1",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] { background-color: #181c26; }
[data-testid="stSidebar"] * { color: #e8eaf0 !important; }
.main-header {
    background: linear-gradient(135deg, #4f8ef7, #7c3aed);
    padding: 18px 24px; border-radius: 12px; margin-bottom: 24px;
    display: flex; align-items: center; gap: 14px;
}
.main-header h1 { color: white; margin: 0; font-size: 22px; }
.main-header p  { color: rgba(255,255,255,0.75); margin: 0; font-size: 13px; }
.stat-box {
    background: #1e2333; border: 1px solid #2a3048;
    border-radius: 12px; padding: 18px 20px; text-align: center;
}
.stat-box .label { font-size: 11px; color: #6b7280; text-transform: uppercase; letter-spacing: 0.6px; }
.stat-box .value { font-size: 28px; font-weight: 800; margin: 6px 0 4px; }
.stat-box .sub   { font-size: 12px; color: #6b7280; }
.ok-badge    { background: rgba(34,197,94,0.15);  color: #22c55e; padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }
.warn-badge  { background: rgba(245,158,11,0.15); color: #f59e0b; padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }
.err-badge   { background: rgba(239,68,68,0.15);  color: #ef4444; padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }
.issue-box {
    background: rgba(239,68,68,0.08); border: 1px solid rgba(239,68,68,0.25);
    border-radius: 10px; padding: 14px 16px; margin-bottom: 10px;
}
.ok-box {
    background: rgba(34,197,94,0.08); border: 1px solid rgba(34,197,94,0.2);
    border-radius: 10px; padding: 12px 16px; margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)

# ─── Session State ─────────────────────────────────────────────────────────────
if "analysis"      not in st.session_state: st.session_state.analysis      = None
if "chat_history"  not in st.session_state: st.session_state.chat_history  = []
if "api_key"       not in st.session_state:
    # Auto-load from Streamlit Cloud secrets if available
    try:    st.session_state.api_key = st.secrets["ANTHROPIC_API_KEY"]
    except: st.session_state.api_key = ""

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt(n):
    try:    return f"${float(n):,.2f}"
    except: return "$0.00"

def extract_pdf_text(file_bytes: bytes) -> str:
    text = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text.append(t)
    return "\n".join(text)

def extract_xlsx_text(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"=== Sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(v) if v is not None else "" for v in row]
            if any(v.strip() for v in row_vals):
                lines.append("\t".join(row_vals))
    return "\n".join(lines)

def status_badge(status: str) -> str:
    if status == "ok":      return '<span class="ok-badge">✓ Match</span>'
    if status == "warning": return '<span class="warn-badge">⚠ Warning</span>'
    return '<span class="err-badge">✗ Error</span>'

# ─── Claude Analysis ──────────────────────────────────────────────────────────
def run_analysis(file_data: list, api_key: str) -> dict:
    """
    file_data: list of {"name": str, "type": "pdf"|"xlsx", "text": str}
    """
    client = anthropic.Anthropic(api_key=api_key)

    files_block = "\n\n".join(
        f"--- FILE: {f['name']} ({f['type'].upper()}) ---\n{f['text']}"
        for f in file_data
    )

    prompt = f"""{files_block}

---
You are a commission analysis agent for MARCTECH2, INC. (Sales Rep: MAR1) working with American Bright Optoelectronics Corp.

IMPORTANT CONTEXT:
- PDF files ("Comm_from_Payment") show REGULAR SALES (Part I) — payments received in a specific month.
- Excel files ("Comm_Report") have TWO parts:
  * Part I (Regular Sales): for the CURRENT month — amount/commission should match the matching PDF.
  * Part II (Distributor Sales): for the PREVIOUS month (one-month lag is intentional and correct).
- Commission rate is consistently 5%.

YOUR TASK — analyze all files and return a JSON object with this EXACT structure:

{{
  "months": [
    {{
      "month": "December 2025",
      "monthKey": "DEC_2025",
      "pdf": {{
        "found": true,
        "fileName": "Comm_from_Payment_MAR1_DEC_.pdf",
        "regularSalesAmount": 6090.00,
        "regularSalesCommission": 304.50,
        "customers": [{{"name": "DELTA CONTROLS, INC.", "invoiceNo": "451014081", "amount": 6090.00, "commission": 304.50}}]
      }},
      "excel": {{
        "found": true,
        "fileName": "COMM_REPORT_MAR1_DEC_.xlsx",
        "part1Amount": 6090.00,
        "part1Commission": 304.50,
        "part2Month": "November 2025",
        "part2Distributors": [{{"name": "INTEGRA", "amount": 572.00, "commission": 28.60}}],
        "part2TotalAmount": 572.00,
        "part2TotalCommission": 28.60,
        "totalCommission": 333.10
      }},
      "reconciliation": {{
        "status": "ok",
        "issues": [],
        "notes": ""
      }}
    }}
  ],
  "summary": {{
    "totalRegularSalesAmount": 0,
    "totalRegularSalesCommission": 0,
    "totalDistributorSalesAmount": 0,
    "totalDistributorSalesCommission": 0,
    "totalCommissionAllMonths": 0,
    "monthsWithIssues": 0,
    "monthsClean": 0
  }},
  "globalIssues": [],
  "insights": "Brief 3-4 sentence narrative about overall commission trends and key observations."
}}

RECONCILIATION RULES:
- status "ok"      → PDF commission matches Excel Part I commission (within $0.01 tolerance)
- status "warning" → minor discrepancy (<$5) OR one file is missing
- status "error"   → significant discrepancy (>=$5) OR clear mismatch
- For each issue: {{"type": "mismatch"|"missing_pdf"|"missing_excel"|"calculation_error", "description": "...", "pdfValue": ..., "excelValue": ...}}

Return ONLY valid JSON. No markdown fences, no explanation.
"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = "".join(b.text for b in response.content if hasattr(b, "text"))
    raw = re.sub(r"```json|```", "", raw).strip()
    return json.loads(raw)


def chat_with_agent(question: str, context: dict, history: list, api_key: str) -> str:
    client = anthropic.Anthropic(api_key=api_key)

    system = (
        "You are a commission analysis assistant for MARCTECH2, INC. (Sales Rep MAR1). "
        "Answer questions about commission data clearly and concisely. Use dollar amounts and be specific.\n\n"
        f"ANALYZED DATA:\n{json.dumps(context, indent=2)}"
    ) if context else (
        "You are a commission analysis assistant. No data has been analyzed yet. "
        "Ask the user to upload files and run analysis first."
    )

    messages = []
    for h in history[-6:]:   # last 3 turns
        messages.append({"role": "user",      "content": h["user"]})
        messages.append({"role": "assistant", "content": h["assistant"]})
    messages.append({"role": "user", "content": question})

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=800,
        system=system,
        messages=messages,
    )
    return "".join(b.text for b in response.content if hasattr(b, "text"))


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    api_key = st.text_input("Anthropic API Key", type="password",
                             value=st.session_state.api_key,
                             placeholder="sk-ant-…")
    if api_key:
        st.session_state.api_key = api_key

    st.markdown("---")
    st.markdown("## 📂 Upload Files")
    st.caption("Upload PDF commission reports and/or Excel statements.")

    uploaded = st.file_uploader(
        "Drag & drop files here",
        type=["pdf", "xlsx", "xls"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded:
        st.markdown(f"**{len(uploaded)} file(s) ready**")
        for f in uploaded:
            icon = "📄" if f.name.endswith(".pdf") else "📊"
            st.caption(f"{icon} {f.name}")

    st.markdown("---")
    run = st.button("⚡ Analyze & Reconcile", use_container_width=True,
                    disabled=not uploaded or not st.session_state.api_key)
    if not st.session_state.api_key:
        st.caption("⚠️ Enter your API key above first.")

# ─── Run Analysis ─────────────────────────────────────────────────────────────
if run and uploaded:
    with st.spinner("Extracting file contents…"):
        file_data = []
        for f in uploaded:
            raw = f.read()
            ext = Path(f.name).suffix.lower()
            if ext == ".pdf":
                text = extract_pdf_text(raw)
                ftype = "pdf"
            else:
                text = extract_xlsx_text(raw)
                ftype = "xlsx"
            file_data.append({"name": f.name, "type": ftype, "text": text})

    with st.spinner("Analyzing with Claude AI…"):
        try:
            st.session_state.analysis = run_analysis(file_data, st.session_state.api_key)
            st.success("Analysis complete!")
        except Exception as e:
            st.error(f"Analysis failed: {e}")

# ─── Main Content ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <div>
    <h1>📊 Commission Intelligence Agent</h1>
    <p>MARCTECH2, INC. · American Bright Optoelectronics Corp. · Sales Rep: MAR1</p>
  </div>
</div>
""", unsafe_allow_html=True)

analysis = st.session_state.analysis

if not analysis:
    st.info("👈 Upload your PDF and Excel commission files in the sidebar, then click **Analyze & Reconcile**.")
    with st.expander("ℹ️ How this works"):
        st.markdown("""
**PDF files** (`Comm_from_Payment_MAR1_XXX.pdf`)
- Show Regular Sales payments received in a given month (Part I)

**Excel files** (`Comm_Report_MAR1_XXX.xlsx`)
- **Part I** — Regular Sales for the *current* month → should match the PDF
- **Part II** — Distributor Sales for the *previous* month (one-month lag is normal)

**The agent will:**
1. Parse all files and extract structured data
2. Cross-check PDF vs Excel Part I amounts for each month
3. Flag any discrepancies, missing files, or calculation errors
4. Generate a consolidated report across all months
5. Let you ask questions via the AI chat tab
        """)
else:
    months   = analysis.get("months", [])
    summary  = analysis.get("summary", {})
    insights = analysis.get("insights", "")

    tab1, tab2, tab3, tab4 = st.tabs(["📊 Summary", "🔍 Reconciliation", "📋 Full Report", "💬 Ask AI"])

    # ── Tab 1: Summary ──────────────────────────────────────────────────────
    with tab1:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Regular Sales", fmt(summary.get("totalRegularSalesAmount", 0)),
                      f"Comm: {fmt(summary.get('totalRegularSalesCommission', 0))}")
        with c2:
            st.metric("Distributor Sales", fmt(summary.get("totalDistributorSalesAmount", 0)),
                      f"Comm: {fmt(summary.get('totalDistributorSalesCommission', 0))}")
        with c3:
            st.metric("Total Commission", fmt(summary.get("totalCommissionAllMonths", 0)),
                      f"{summary.get('monthsClean', 0)} clean months")
        with c4:
            issues = summary.get("monthsWithIssues", 0)
            st.metric("Issues Found", issues,
                      delta="All clear ✓" if issues == 0 else f"{issues} month(s) need review",
                      delta_color="normal" if issues == 0 else "inverse")

        if insights:
            st.markdown("### 🤖 AI Insights")
            st.info(insights)

        st.markdown("### Monthly Breakdown")
        import pandas as pd
        rows = []
        for m in months:
            rows.append({
                "Month":           m["month"],
                "Regular Sales":   m.get("excel", {}).get("part1Amount") or m.get("pdf", {}).get("regularSalesAmount", 0),
                "Reg. Commission": m.get("excel", {}).get("part1Commission") or m.get("pdf", {}).get("regularSalesCommission", 0),
                "Disty Sales":     m.get("excel", {}).get("part2TotalAmount", 0),
                "Disty Comm.":     m.get("excel", {}).get("part2TotalCommission", 0),
                "Total Comm.":     m.get("excel", {}).get("totalCommission") or m.get("pdf", {}).get("regularSalesCommission", 0),
                "Status":          m.get("reconciliation", {}).get("status", "").upper(),
            })
        df = pd.DataFrame(rows)
        for col in ["Regular Sales", "Reg. Commission", "Disty Sales", "Disty Comm.", "Total Comm."]:
            df[col] = df[col].apply(lambda x: f"${x:,.2f}")
        st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Tab 2: Reconciliation ───────────────────────────────────────────────
    with tab2:
        all_ok = all(m.get("reconciliation", {}).get("status") == "ok" for m in months)
        if all_ok:
            st.success("✅ All months reconciled perfectly — no discrepancies found.")

        for m in months:
            rec = m.get("reconciliation", {})
            status = rec.get("status", "ok")
            issues = rec.get("issues", [])

            if status == "ok":
                pdf_comm = m.get("pdf", {}).get("regularSalesCommission", 0)
                st.markdown(
                    f'<div class="ok-box">✅ <strong>{m["month"]}</strong> — '
                    f'PDF and Excel match. Commission: {fmt(pdf_comm)}</div>',
                    unsafe_allow_html=True,
                )
            else:
                for issue in issues:
                    desc = issue.get("description", "")
                    pv   = issue.get("pdfValue")
                    xv   = issue.get("excelValue")
                    detail = f"<br><b>PDF:</b> {fmt(pv)}  &nbsp;·&nbsp;  <b>Excel:</b> {fmt(xv)}" if pv is not None else ""
                    st.markdown(
                        f'<div class="issue-box"><strong>⚠️ {m["month"]}</strong><br>{desc}{detail}</div>',
                        unsafe_allow_html=True,
                    )

        st.markdown("### Detail Table")
        rec_rows = []
        for m in months:
            pdf_comm = m.get("pdf",   {}).get("regularSalesCommission")
            xl_comm  = m.get("excel", {}).get("part1Commission")
            diff     = abs(pdf_comm - xl_comm) if (pdf_comm is not None and xl_comm is not None) else None
            rec_rows.append({
                "Month":          m["month"],
                "PDF Found":      "✓" if m.get("pdf",   {}).get("found") else "✗",
                "Excel Found":    "✓" if m.get("excel", {}).get("found") else "✗",
                "PDF Comm.":      fmt(pdf_comm)  if pdf_comm is not None else "—",
                "Excel P1 Comm.": fmt(xl_comm)   if xl_comm  is not None else "—",
                "Difference":     fmt(diff)       if diff is not None else "—",
                "Result":         m.get("reconciliation", {}).get("status", "").upper(),
            })
        st.dataframe(pd.DataFrame(rec_rows), use_container_width=True, hide_index=True)

    # ── Tab 3: Full Report ──────────────────────────────────────────────────
    with tab3:
        for m in months:
            with st.expander(f"📅 {m['month']}", expanded=True):
                # Part I
                st.markdown("**Part I — Regular Sales**")
                customers = m.get("pdf", {}).get("customers", [])
                if customers:
                    cdf = pd.DataFrame(customers)
                    for col in ["amount", "commission"]:
                        if col in cdf.columns:
                            cdf[col] = cdf[col].apply(lambda x: f"${x:,.2f}")
                    st.dataframe(cdf, use_container_width=True, hide_index=True)
                else:
                    st.caption("No PDF data available.")

                p1a = m.get("excel", {}).get("part1Amount") or m.get("pdf", {}).get("regularSalesAmount", 0)
                p1c = m.get("excel", {}).get("part1Commission") or m.get("pdf", {}).get("regularSalesCommission", 0)
                st.caption(f"Part I Total: {fmt(p1a)}  |  Commission: {fmt(p1c)}")

                # Part II
                distys = m.get("excel", {}).get("part2Distributors", [])
                if distys:
                    p2month = m.get("excel", {}).get("part2Month", "prior month")
                    st.markdown(f"**Part II — Distributor Sales** *(for {p2month})*")
                    ddf = pd.DataFrame(distys)
                    for col in ["amount", "commission"]:
                        if col in ddf.columns:
                            ddf[col] = ddf[col].apply(lambda x: f"${x:,.2f}")
                    st.dataframe(ddf, use_container_width=True, hide_index=True)
                    p2a = m.get("excel", {}).get("part2TotalAmount", 0)
                    p2c = m.get("excel", {}).get("part2TotalCommission", 0)
                    st.caption(f"Part II Total: {fmt(p2a)}  |  Commission: {fmt(p2c)}")

                total = m.get("excel", {}).get("totalCommission") or m.get("pdf", {}).get("regularSalesCommission", 0)
                st.success(f"**Total Commission for {m['month']}: {fmt(total)}**")

        st.markdown("---")
        grand = summary.get("totalCommissionAllMonths", 0)
        st.markdown(f"## 💰 Grand Total Commission: {fmt(grand)}")

    # ── Tab 4: Ask AI ───────────────────────────────────────────────────────
    with tab4:
        st.markdown("Ask anything about your commission data.")

        # Display history
        for h in st.session_state.chat_history:
            with st.chat_message("user"):
                st.write(h["user"])
            with st.chat_message("assistant"):
                st.write(h["assistant"])

        question = st.chat_input("Ask about commissions, discrepancies, totals…")
        if question:
            if not st.session_state.api_key:
                st.error("Enter your Anthropic API key in the sidebar.")
            else:
                with st.chat_message("user"):
                    st.write(question)
                with st.chat_message("assistant"):
                    with st.spinner("Thinking…"):
                        reply = chat_with_agent(
                            question,
                            st.session_state.analysis,
                            st.session_state.chat_history,
                            st.session_state.api_key,
                        )
                    st.write(reply)
                st.session_state.chat_history.append({"user": question, "assistant": reply})
